from __future__ import annotations

"""Excel export utilities."""

from pathlib import Path

from openpyxl import Workbook

from .models import InvoiceRecord


HEADERS = ["Invoice #", "Date", "File Name", "Status"]


def export_records_to_excel(records: list[InvoiceRecord], output_path: Path, sheet_name: str) -> Path:
    """Export processed invoice records to an Excel workbook.

    Args:
        records: Records produced by the processing pipeline.
        output_path: Destination XLSX path.
        sheet_name: Worksheet title to use inside the workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for index, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=index, value=header)

    for row_index, record in enumerate(records, start=2):
        ws.cell(row=row_index, column=1, value=record.invoice_number)
        ws.cell(row=row_index, column=2, value=record.invoice_date)
        ws.cell(row=row_index, column=3, value=record.file_name)
        ws.cell(row=row_index, column=4, value=record.status)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
